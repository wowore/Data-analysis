import os
import openpyxl
import pandas as pd


class dataSet:
    def __init__(self):
        self.chinaDict = {}
        self.worldDict = {}
        self.province = ['省份', '累计确诊', '死亡', '治愈', '现有确诊', '累计确诊增量', '死亡增量', '治愈增量', '现有确诊增量']
        self.country = ['国家', '累计确诊', '死亡', '治愈', '现有确诊', '累计确诊增量']
        self.xlsxList = [file for file in os.listdir('Data') if '.xlsx' in file]

    def initDicts(self):
        pass

    def loadXlsx(self):
        for xlsxName in self.xlsxList:
            time = xlsxName[4:-5]
            wb = openpyxl.load_workbook('Data/' + xlsxName)
            shName = wb.sheetnames
            # chinaSheet = wb['国内疫情']
            # print(chinaSheet.max_column, chinaSheet.max_row)
            # print(time)
            x = pd.ExcelFile(r'Data/' + xlsxName)
            for name in shName[:1]:
                pro = '省份'
                print(f'{name}>>>>>>>\n{list(pd.read_excel(x, name)[pro])}')


if __name__ == '__main__':
    ds = dataSet()
    print(ds.xlsxList)
    ds.loadXlsx()
