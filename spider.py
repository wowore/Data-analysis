import time

import requests
from lxml import etree
import json
import re, os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options  # 无界面浏览器


class GetData:
    def __init__(self):
        self.globalDataDict = {}
        self.ChinaData = {}
        self.ChinaConfirmed = 0
        self.ChinaDied = 0
        self.ChinaCured = 0
        self.ChinaCurConfirm = 0
        self.ChinaConfirmedRelative = 0
        self.htmlFile = ''
        self.ChinaHtmlFile = ''
        self.WorldHtmlFile = ''
        self.flag = False
        self.time = ''
        self.time1, self.time2 = self.getTime()

    def getNews(self):
        ChinaContentDict = {'head': '国内最新疫情资讯'}
        WorldContentDict = {'head': '国外最新疫情资讯'}
        with open('baseHtml.html', 'r', encoding='utf-8') as f:
            self.htmlFile = f.read()
        f.close()
        china = self.htmlFile
        world = self.htmlFile
        baseUrl = "https://voice.baidu.com/act/newpneumonia/newpneumonia/"
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        driver = webdriver.Chrome(executable_path='F:/综合课程设计III/数据分析/chromedriver.exe',
                                  chrome_options=chrome_options)
        driver.get(baseUrl)
        resp_text = driver.page_source
        page_html = etree.HTML(resp_text)
        a_list = page_html.xpath('//*/div[@class="Virus_1-1-350_TB6x3k"]/a')
        i = 1
        for a in a_list:
            url = a.xpath('./@href')[0]
            div = a.xpath('./div')
            title = div[0].xpath('./text()')[0]
            ChinaContentDict[f'href{i}'] = url
            ChinaContentDict[f'title{i}'] = title
            i += 1
        driver.find_element_by_xpath('//*[@id="ptab-1"]/div[2]/div[2]').click()
        time.sleep(3)
        # driver.get(baseUrl)
        resp_text = driver.page_source
        page_html = etree.HTML(resp_text)
        b_list = page_html.xpath('//*/div[@class="Virus_1-1-350_TB6x3k"]/a')
        i = 1
        for b in b_list:
            url = b.xpath('./@href')[0]
            div = b.xpath('./div')
            title = div[0].xpath('./text()')[0]
            WorldContentDict[f'href{i}'] = url
            WorldContentDict[f'title{i}'] = title
            i += 1
        driver.quit()

        self.ChinaHtmlFile = china.format_map(ChinaContentDict)
        self.WorldHtmlFile = world.format_map(WorldContentDict)
        with open('ChinaNews.html', 'w', encoding='utf-8') as f:
            f.write(self.ChinaHtmlFile)
        f.close()
        with open('WorldNews.html', 'w', encoding='utf-8') as f:
            f.write(self.WorldHtmlFile)
        f.close()

    def getData(self):
        # 目标url
        url = "https://voice.baidu.com/act/newpneumonia/newpneumonia/"

        # 伪装请求头
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/80.0.3987.149 Safari/537.36 '
        }

        # 发出get请求
        response = requests.get(url, headers=headers)

        # 将请求的结果写入文件,便于分析
        with open('html.txt', 'w') as file:
            file.write(response.text)

    def getTime(self):
        with open('html.txt', 'r') as file:
            text = file.read()
        # 获取更新时间
        time_in = re.findall('"mapLastUpdatedTime":"(.*?)"', text)[0]
        time_out = re.findall('"foreignLastUpdatedTime":"(.*?)"', text)[0]
        print('国内疫情更新时间为 ' + time_in)
        print('国外疫情更新时间为 ' + time_out)
        return time_in, time_out

    def parseData(self, time):
        with open('html.txt', 'r') as file:
            text = file.read()
        # 生成HTML对象
        html = etree.HTML(text)
        # 解析数据
        result = html.xpath('//script[@type="application/json"]/text()')
        result = result[0]
        result = json.loads(result)
        result = json.dumps(result['component'][0]['caseList'])
        with open(f'Data/Data{time[:10]}.json', 'w') as file:
            file.write(result)
            print('数据已写入json文件...')

        response = requests.get("https://voice.baidu.com/act/newpneumonia/newpneumonia/")
        # 将请求的结果写入文件,便于分析
        with open('html.txt', 'w') as file:
            file.write(response.text)

        # 获取时间
        time_in = re.findall('"mapLastUpdatedTime":"(.*?)"', response.text)[0]
        time_out = re.findall('"foreignLastUpdatedTime":"(.*?)"', response.text)[0]
        print(time_in)
        print(time_out)

        # 生成HTML对象
        html = etree.HTML(response.text)
        # 解析数据
        result = html.xpath('//script[@type="application/json"]/text()')
        result = result[0]
        result = json.loads(result)
        # 以每个省的数据为一个字典
        data_in = result['component'][0]['caseList']
        data_out = result['component'][0]['globalList']

        '''
        area --> 大多为省份
        city --> 城市
        confirmed --> 累计
        died --> 死亡
        crued --> 治愈
        relativeTime --> 
        confirmedRelative --> 累计的增量
        curedRelative --> 治愈的增量
        curConfirm --> 现有确诊
        curConfirmRelative --> 现有确诊的增量
        diedRelative --> 死亡的增量
        '''

        # 规律----遍历列表的每一项,可以发现,每一项(type:字典)均代表一个省份等区域,这个字典的前11项是该省份的疫情数据,
        # 当key = 'subList'时,其结果为只有一项的列表,提取出列表的第一项,得到一系列的字典,字典中包含该城市的疫情数据.

        # 将得到的数据写入excel文件
        # 创建一个工作簿
        wb = openpyxl.Workbook()
        # 创建工作表,每一个工作表代表一个area
        ws_in = wb.active
        ws_in.title = "国内疫情"
        ws_in.append(['省份', '累计确诊', '死亡', '治愈', '现有确诊', '累计确诊增量', '死亡增量', '治愈增量', '现有确诊增量'])
        for each in data_in:
            temp_list = [each['area'], each['confirmed'], each['died'], each['crued'], each['curConfirm'],
                         each['confirmedRelative'], each['diedRelative'], each['curedRelative'],
                         each['curConfirmRelative']]
            self.ChinaData[each['area']] = [each['confirmed'], each['died'], each['crued']]
            self.ChinaConfirmed += int(each['confirmed'])
            self.ChinaDied += int(each['died'])
            self.ChinaCured += int(each['crued'])
            self.ChinaCurConfirm += int(each['curConfirm'])
            self.ChinaConfirmedRelative += int(each['confirmedRelative'])

            for i in range(len(temp_list)):
                if temp_list[i] == '':
                    temp_list[i] = '0'
            ws_in.append(temp_list)
        list_temp_china = ['中国', str(self.ChinaConfirmed), str(self.ChinaDied), str(self.ChinaCured),
                           str(self.ChinaCurConfirm), str(self.ChinaConfirmedRelative)]
        ws_out = wb.create_sheet('各国疫情')
        ws_out.append(['国家', '累计确诊', '死亡', '治愈', '现有确诊', '累计确诊增量'])
        ws_out.append(list_temp_china)
        # 获取国外疫情数据
        for each in data_out:
            # 创建一个新的工作表
            # ws_out = wb.create_sheet(sheet_title)
            for country in each['subList']:
                if country == '钻石公主号邮轮':
                    continue
                self.globalDataDict[country['country']] = [country['confirmed'], country['died'], country['crued']]
                list_temp = [country['country'], country['confirmed'], country['died'], country['crued'],
                             country['curConfirm'], country['confirmedRelative']]
                for i in range(len(list_temp)):
                    if list_temp[i] == '':
                        list_temp[i] = '0'
                ws_out.append(list_temp)

            # 保存excel文件
        wb.save(f'Data/Data{time[:10]}.xlsx')

    def runSpider(self):
        self.time = self.time1[:10]
        # flag = True
        # fileList = os.listdir('./Data')
        # for file in fileList:
        #     if self.time1[:10] in file:
        #         flag = False
        # if flag:
        self.getNews()
        self.getData()
        self.parseData(self.time1)
        self.globalDataDict['中国'] = [str(self.ChinaConfirmed), str(self.ChinaDied), str(self.ChinaCured)]
        with open(f'Data/globalData{self.time2[:10]}.json', 'w') as f:
            json.dump(self.globalDataDict, f, ensure_ascii=False)
        with open(f'Data/chinaData{self.time1[:10]}.json', 'w') as f:
            json.dump(self.ChinaData, f, indent=2, ensure_ascii=False)
        self.flag = True

#
# if __name__ == '__main__':
#     gd = GetData()
#     gd.getNews()
